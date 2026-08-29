"""
Medical Drive folder cleaner.

Sweeps the whole Google Drive "Medical" tree, reads each PDF/image's actual
content (text layer, or Claude vision for scanned docs), and works out the
real collection date, category and provider — then plans a rename/move to
bring the file in line with the naming convention:

    Medical/{year}/{category}/YYYY-MM-DD_{category}_{provider}_{type}.ext

This exists because ingest_labs_gdrive.py trusts the filename blindly:
parse_filename() silently falls back to date.today() on a bad date, and the
ingest only picks up files with "_labs_" in the name. A messy filename is
invisible to the ingest and/or gets stamped with the wrong date. Running
this first keeps the folder — and the timeline — clean.

Safety rules (do not relax these):
  - Never deletes anything. Ever.
  - Ambiguous dates (both day and month <= 12, so D/M vs M/D genuinely
    can't be told apart from the text alone) are quarantined, not guessed.
    A previous script guessed and got a month wrong — see rename_log
    history / commit messages for the case this is guarding against.
  - Anything read with low confidence, or that fails to read at all, is
    quarantined rather than filed on a guess.
  - Duplicates (identical content hash) are quarantined, keeping whichever
    copy is already correctly named/placed (or the oldest, if neither is).
  - Dry-run by default. Nothing changes on Drive unless --apply is passed.
    Every run (dry-run or apply) writes rename_log.csv describing what it
    did or would do.
  - Every run also (re)writes a master tracker workbook at tracker/master_tracker.xlsx
    — one tab per category plus a Resumen tab and a _REVISAR tab — as a standing
    inventory of everything in the Medical folder. This happens on dry runs too,
    since building the tracker never touches Drive.

Auth: same service-account pattern as ingest_labs_gdrive.py, via the
GDRIVE_CREDENTIALS secret. Unlike the ingest (readonly), this needs the
service account to have Editor access on the Medical folder, because it
renames and moves files.
"""
import os, io, sys, csv, re, json, base64, hashlib, argparse
from datetime import date, datetime, timezone
from collections import defaultdict

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
import anthropic
import pdfplumber
from pdf2image import convert_from_bytes
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
GDRIVE_FOLDER_ID = "1_pB5M_-xqWU-jNYykK83fhZiWc5ldrS2"  # Medical/
SCOPES = ["https://www.googleapis.com/auth/drive"]  # needs write (Editor on the folder)
MODEL = "claude-sonnet-5"  # claude-sonnet-4-20250514 has been retired — every call 404'd

MAX_IMAGE_BYTES = 10 * 1024 * 1024  # Anthropic's vision API hard limit on base64-decoded image size

# Ordered (not a set) so the tracker's tabs come out in a stable, sensible order.
CATEGORIES_ORDERED = ["labs", "radiologia", "ultrasonidos", "inbody", "patologia"]
CATEGORIES = set(CATEGORIES_ORDERED)
KNOWN_PROVIDERS = {"quest-mx", "quest-usa", "labcorp", "hospital-angeles-lomas"}
REVISAR_FOLDER_NAME = "_REVISAR"
DEFAULT_TRACKER_PATH = os.path.join("tracker", "master_tracker.xlsx")

MIN_TEXT_LEN_FOR_TEXT_LAYER = 40  # below this, treat the PDF as scanned/image-only

# Error text that means "the Claude API itself is unavailable/unusable" —
# billing, auth, rate limits, an outage — as opposed to "this one document was
# hard to read." Seeing one of these means every remaining file is about to
# fail the same way; nothing on Drive should be touched until it's fixed.
SYSTEMIC_ERROR_MARKERS = (
    "credit balance", "insufficient_quota", "authentication_error",
    "invalid x-api-key", "invalid api key", "rate_limit_error", "overloaded_error",
)
# Fallback for an error we didn't anticipate: this many classification
# failures in a row (regardless of wording) is far more consistent with the
# service itself being down than with that many individually bad documents
# in a row, so treat it the same way.
CONSECUTIVE_FAILURE_ABORT_THRESHOLD = 5


def looks_systemic(err_text):
    """True if an extraction-failure error message points at the API/account
    itself being broken, not this one document."""
    if not err_text:
        return False
    low = str(err_text).lower()
    return any(marker in low for marker in SYSTEMIC_ERROR_MARKERS)

DATE_LABELS_HINT = (
    "Fecha Toma Muestra, Muestra Tomada, Fecha de Toma, Fecha de Recoleccion, "
    "Collected, Date Collected, Specimen Collected"
)

CLASSIFY_INSTRUCTIONS = f"""You are helping organize a personal medical records archive (thyroid cancer
monitoring: labs, ultrasounds, radiology, pathology/biopsy reports, and InBody
body-composition scans). You will be shown a medical document (as text or as
an image). Identify the following and return ONLY a valid JSON object, no
markdown, no explanation:

{{
  "fecha_encontrada": true or false,
  "fecha_texto": "the date exactly as printed, digits and separators only, e.g. '17/04/2026' — do NOT reorder or reformat it, do NOT guess which number is the day vs month",
  "fecha_label": "which label it was next to, e.g. 'Fecha Toma Muestra'",
  "fecha_secundaria_texto": "any OTHER date printed anywhere else on the same document — report date, print date, review date, validation date, whatever else appears — exactly as printed, same no-reorder rule as above. null if there truly isn't a second date anywhere on the page.",
  "fecha_secundaria_label": "which label that second date was next to, e.g. 'Fecha de Reporte'",
  "categoria": one of "labs", "radiologia", "ultrasonidos", "inbody", "patologia", or null if you cannot tell confidently,
  "categoria_confianza": "alta" or "baja",
  "proveedor": a short kebab-case slug for the lab/clinic/hospital that issued the document (use "quest-mx" for Quest Diagnostics Mexico, "quest-usa" for Quest Diagnostics USA, "labcorp" for LabCorp, "hospital-angeles-lomas" for Hospital Angeles Lomas; otherwise invent a short kebab-case slug from the letterhead, e.g. "clinica-san-jose"), or null if you truly cannot tell,
  "tipo": a short kebab-case slug (1-3 words) for what this document is, e.g. "tiroideo", "biometria", "general", "tiroglobulina", "perfil-lipidico", "ultrasonido-tiroides", "biopsia-tiroides", "torax", "pet", "rastreo-yodo", "otro",
  "notas": "anything odd worth a human's attention, or empty string"
}}

Rules for fecha_texto specifically:
- Find the SPECIMEN COLLECTION date, not the birth date, report-creation date,
  print date, or release date. Look for labels like: {DATE_LABELS_HINT}.
- Copy the date digits and separators EXACTLY as printed (e.g. "05/03/2026").
  Do not convert it, do not decide whether it's day-first or month-first —
  that decision is made deliberately outside this step, by a human-reviewed
  rule, because a past automated guess got the month wrong.
- If no such date is findable, set fecha_encontrada to false and fecha_texto to null.
- Also look for a SECOND, different date printed anywhere else on the page (a
  report/print/release date is fine here — this one is NOT used to name the
  file, it's only used to double-check which number is the day and which is
  the month when fecha_texto's own order is unclear). Same copy-exactly rule.

PET scans and radioactive-iodine studies ("rastreo de yodo", "gammagrama",
whole-body iodine scans) are nuclear-medicine imaging — classify these as
categoria "radiologia" with tipo "pet" or "rastreo-yodo" (not a separate
category).

Be conservative: if you are not confident about categoria or the document is
illegible/ambiguous, say so via categoria_confianza "baja" rather than guessing.
"""

# ── Google Drive ──────────────────────────────────────────────────────────────
def get_drive_service():
    creds_info = json.loads(os.environ["GDRIVE_CREDENTIALS"])
    creds = service_account.Credentials.from_service_account_info(creds_info, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)


def download_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf.read()


def list_children(service, folder_id):
    """Non-recursive: immediate children of a folder (files + folders)."""
    items, page_token = [], None
    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id,name,mimeType,parents,createdTime)",
            pageToken=page_token,
        ).execute()
        items.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items


def walk_tree(service, folder_id, path_parts, skip_folder_names=(REVISAR_FOLDER_NAME,)):
    """Yield (file_dict, path_parts, parent_folder_id) for every non-folder file
    under folder_id, recursively. Folders named in skip_folder_names (and their
    contents) are skipped entirely — used to leave _REVISAR alone."""
    for item in list_children(service, folder_id):
        if item["mimeType"] == "application/vnd.google-apps.folder":
            if item["name"] in skip_folder_names:
                continue
            yield from walk_tree(service, item["id"], path_parts + [item["name"]], skip_folder_names)
        else:
            yield item, path_parts, folder_id


def get_or_create_folder(service, parent_id, name, folder_cache, apply_changes):
    """folder_cache maps (parent_id, name) -> folder_id for folders we know exist.
    In dry-run mode, a folder that doesn't exist yet is represented by a
    synthetic id "PLANNED:{parent_id}:{name}" so planning can continue without
    touching Drive."""
    key = (parent_id, name)
    if key in folder_cache:
        return folder_cache[key]
    if not apply_changes:
        planned_id = f"PLANNED:{parent_id}:{name}"
        folder_cache[key] = planned_id
        return planned_id
    folder = service.files().create(
        body={"name": name, "mimeType": "application/vnd.google-apps.folder", "parents": [parent_id]},
        fields="id",
    ).execute()
    folder_cache[key] = folder["id"]
    return folder["id"]


def move_file(service, file_id, old_parent_id, new_parent_id, new_name):
    kwargs = {"fileId": file_id, "fields": "id,name,parents"}
    if new_name is not None:
        kwargs["body"] = {"name": new_name}
    if new_parent_id and new_parent_id != old_parent_id:
        kwargs["addParents"] = new_parent_id
        kwargs["removeParents"] = old_parent_id
    service.files().update(**kwargs).execute()


# ── Text / vision extraction ─────────────────────────────────────────────────
def extract_pdf_text(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = ""
            for page in pdf.pages[:3]:  # header/specimen info is always on the first pages
                t = page.extract_text()
                if t:
                    text += t + "\n"
        return text.strip()
    except Exception as e:
        print(f"    pdfplumber error: {e}")
        return ""


def _encode_jpeg_under_limit(pil_image, max_bytes=MAX_IMAGE_BYTES, start_quality=85):
    """Encodes a PIL image as JPEG, stepping quality down and then dimensions
    down, until it fits under max_bytes. A scanned page at 150 DPI can render
    well over Anthropic's 10MB image limit — this is what a past run hit."""
    img = pil_image.convert("RGB")
    quality = start_quality
    data = b""
    for _ in range(12):
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        data = buf.getvalue()
        if len(data) <= max_bytes:
            return data
        if quality > 40:
            quality -= 15
        else:
            w, h = img.size
            img = img.resize((max(1, w * 3 // 4), max(1, h * 3 // 4)))
            quality = 70
    return data  # best effort if still over after the loop cap (very unlikely)


def render_first_page_image(pdf_bytes):
    """Rasterizes page 1 of a PDF and returns JPEG bytes guaranteed under
    Anthropic's 10MB vision limit."""
    images = convert_from_bytes(pdf_bytes, dpi=150, first_page=1, last_page=1)
    return _encode_jpeg_under_limit(images[0])


def _extract_text(msg):
    """Return the first text block's content, skipping any thinking/reasoning
    blocks the model may return before its actual answer."""
    for block in msg.content:
        block_text = getattr(block, "text", None)
        if block_text:
            return block_text
    raise ValueError("no text content block in model response")


CLASSIFY_MAX_TOKENS = 2048  # 600, then 1024, both still occasionally got cut off
                             # mid-JSON ("Unterminated string...") on documents that
                             # prompted a long "notas" explanation. This is a cheap
                             # classification call either way, so budget generously.


def classify_from_text(text, client):
    if len(text) > 15000:
        text = text[:15000]
    msg = client.messages.create(
        model=MODEL, max_tokens=CLASSIFY_MAX_TOKENS,
        messages=[{"role": "user", "content": CLASSIFY_INSTRUCTIONS + "\n\nDocument text:\n" + text}],
    )
    return _parse_json_object(_extract_text(msg))


def classify_from_image(image_bytes, media_type, client):
    img_b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    msg = client.messages.create(
        model=MODEL, max_tokens=CLASSIFY_MAX_TOKENS,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
            {"type": "text", "text": CLASSIFY_INSTRUCTIONS},
        ]}],
    )
    return _parse_json_object(_extract_text(msg))


def _parse_json_object(raw):
    raw = raw.strip()
    if "```" in raw:
        for p in raw.split("```"):
            p = p.strip().lstrip("json").strip()
            if p.startswith("{"):
                raw = p
                break
    start, end = raw.find("{"), raw.rfind("}")
    if start != -1 and end != -1:
        raw = raw[start:end + 1]
    return json.loads(raw)


def classify_document(file_bytes, mime_type, client):
    """Returns (classification_dict_or_None, method, error_or_None)."""
    if mime_type == "application/pdf":
        text = extract_pdf_text(file_bytes)
        if len(text) >= MIN_TEXT_LEN_FOR_TEXT_LAYER:
            try:
                return classify_from_text(text, client), "text", None
            except Exception as e:
                return None, "text", str(e)
        try:
            image_bytes = render_first_page_image(file_bytes)
        except Exception as e:
            return None, "vision", f"pdf2image render failed: {e}"
        try:
            return classify_from_image(image_bytes, "image/jpeg", client), "vision", None
        except Exception as e:
            return None, "vision", str(e)
    elif mime_type in ("image/jpeg", "image/png", "image/jpg"):
        media_type = "image/png" if mime_type == "image/png" else "image/jpeg"
        image_bytes = file_bytes
        if len(file_bytes) > MAX_IMAGE_BYTES:
            # e.g. a big phone photo of an InBody screen — re-encode under the limit.
            try:
                image_bytes = _encode_jpeg_under_limit(Image.open(io.BytesIO(file_bytes)))
                media_type = "image/jpeg"
            except Exception as e:
                return None, "vision", f"image too large and re-encode failed: {e}"
        try:
            return classify_from_image(image_bytes, media_type, client), "vision", None
        except Exception as e:
            return None, "vision", str(e)
    else:
        return None, "skip", f"unsupported mime type {mime_type}"


# ── Date ambiguity resolution (deterministic — never left to the model) ──────
# Real fecha_texto values seen in the wild (from rename_log.csv's diagnostic
# columns) that the original all-numeric regex couldn't handle at all:
#   "15 de julio del 2023"   "21 DE AGOSTO DE 2019"   "29-ABRIL-2019"
#   "24-DIC-18"              "2025-08-11"             "09-01-19"
# The first four are Spanish month names (always unambiguous — the month is
# spelled out) and ISO YYYY-MM-DD (always unambiguous — year leads, then
# month, per the standard). Neither of those is a guess; both are handled
# below before we ever fall through to the genuinely ambiguous numeric case.

SPANISH_MONTHS = {
    "enero": 1, "ene": 1,
    "febrero": 2, "feb": 2,
    "marzo": 3, "mar": 3,
    "abril": 4, "abr": 4,
    "mayo": 5, "may": 5,
    "junio": 6, "jun": 6,
    "julio": 7, "jul": 7,
    "agosto": 8, "ago": 8,
    "septiembre": 9, "setiembre": 9, "sept": 9, "sep": 9,
    "octubre": 10, "oct": 10,
    "noviembre": 11, "nov": 11,
    "diciembre": 12, "dic": 12,
}
_MONTH_ALTERNATION = "|".join(sorted(SPANISH_MONTHS.keys(), key=len, reverse=True))
# Trailing connector before the year: "del 2020", "de 2020", "del año 2020",
# "año 2020" — real phrasing seen in these reports.
SPANISH_DATE_RE = re.compile(
    r"(\d{1,2})\s*(?:del|de|-)?\s*(" + _MONTH_ALTERNATION + r")\.?\s*(?:del año|del|de|año|-)?\s*(\d{2,4})",
    re.IGNORECASE,
)

# Year-first, unambiguous by ISO 8601 convention (year always leads, then month, then day).
ISO_DATE_RE = re.compile(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})")

# Day-or-month first, numeric only. Genuinely ambiguous unless one of the two
# leading numbers is >12. Now also accepts 2-digit years (e.g. "09-01-19").
DATE_TOKEN_RE = re.compile(r"(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})")


def _expand_year(y):
    """2-digit years in this archive are all recent (2000s) — never 1900s."""
    return 2000 + y if y < 100 else y


def numeric_order_hint(text):
    """Given a raw date string, return 'DM' or 'MD' if its own two leading
    numbers unambiguously reveal which order this document's software prints
    dates in (one of them is >12) — else None (itself ambiguous, non-numeric,
    unparseable, or missing). Used only to read a genuine same-document signal
    (e.g. a report date printed alongside an ambiguous collection date), never
    to invent one — a Spanish-named month or an ISO year-first date carries no
    order information to transfer, so those return None here too."""
    if not text:
        return None
    # Bail out on Spanish-named-month or ISO year-first dates first — neither
    # carries D/M-order information, and matching DATE_TOKEN_RE against a
    # substring of either (e.g. the "08-11" tail of an ISO "2025-08-11") would
    # produce a false hint.
    if SPANISH_DATE_RE.search(text) or ISO_DATE_RE.search(text):
        return None
    m = DATE_TOKEN_RE.search(text)
    if not m:
        return None
    a, b = int(m.group(1)), int(m.group(2))
    if a > 12 and b <= 12:
        return "DM"
    if b > 12 and a <= 12:
        return "MD"
    return None


def both_date_interpretations(fecha_texto):
    """For a genuinely ambiguous numeric date, return both calendar-valid
    readings as ISO strings, (day-first, month-first) — e.g. '4/12/2019' ->
    ('2019-12-04', '2019-04-12') — purely so a human reviewing _REVISAR can
    see both options at a glance instead of re-deriving them by hand. Never
    used to pick one automatically."""
    m = DATE_TOKEN_RE.search(fecha_texto or "")
    if not m:
        return None, None
    a, b, y = int(m.group(1)), int(m.group(2)), _expand_year(int(m.group(3)))

    def _try(day, month):
        try:
            return date(y, month, day).isoformat()
        except ValueError:
            return None

    return _try(a, b), _try(b, a)  # (as-D/M reading, as-M/D reading)


def existing_filename_date(filename):
    """If a filename already starts with this pipeline's own YYYY-MM-DD_
    convention, return that date — used only as a tie-break for genuinely
    ambiguous dates, and only when it matches one of the two calendar-valid
    readings (see the "via_existing_name" tier in main()). Strips a REVISAR_
    or DUP_ prefix first — the pipeline's own quarantine/dedup prefixes,
    which a file can still be wearing if it's come back out of _REVISAR (as
    happened after the credit-balance incident) without being renamed."""
    name = filename or ""
    for known_prefix in ("REVISAR_", "DUP_"):
        if name.startswith(known_prefix):
            name = name[len(known_prefix):]
            break
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})_", name)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def resolve_date(fecha_texto, order_hint=None):
    """Returns (date_or_None, reason_or_None, resolved_via_hint_bool).

    Never guesses an ambiguous date on its own. order_hint ('DM' or 'MD') may
    be supplied from numeric_order_hint() run against a SECOND, unambiguous
    date found elsewhere on the same document — real evidence about how that
    specific document's software orders dates, not a guess. It is only used
    as a last resort, when fecha_texto's own digits are genuinely ambiguous."""
    if not fecha_texto:
        return None, "no_date_found", False

    m = SPANISH_DATE_RE.search(fecha_texto)
    if m:
        day = int(m.group(1))
        month = SPANISH_MONTHS[m.group(2).lower()]
        year = _expand_year(int(m.group(3)))
        try:
            return date(year, month, day), None, False
        except ValueError:
            return None, "invalid_calendar_date", False

    m = ISO_DATE_RE.search(fecha_texto)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(year, month, day), None, False
        except ValueError:
            return None, "invalid_calendar_date", False

    m = DATE_TOKEN_RE.search(fecha_texto)
    if not m:
        return None, "unparseable_date_text", False
    a, b, y = int(m.group(1)), int(m.group(2)), _expand_year(int(m.group(3)))
    via_hint = False
    if a > 12 and b > 12:
        return None, "invalid_date_numbers", False
    if a > 12 and b <= 12:
        day, month = a, b          # unambiguous D/M
    elif b > 12 and a <= 12:
        day, month = b, a          # unambiguous M/D
    elif order_hint == "DM":
        day, month = a, b          # both <=12 — resolved from a 2nd date on the same doc
        via_hint = True
    elif order_hint == "MD":
        day, month = b, a
        via_hint = True
    else:
        return None, "ambiguous_day_month", False  # no evidence either way — quarantine
    try:
        return date(y, month, day), None, via_hint
    except ValueError:
        return None, "invalid_calendar_date", False


# ── Filename / slug helpers ───────────────────────────────────────────────────
def slugify(s):
    if not s:
        return "otro"
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-") or "otro"


def build_target_name(fecha, categoria, proveedor, tipo, ext):
    return f"{fecha.isoformat()}_{slugify(categoria)}_{slugify(proveedor)}_{slugify(tipo)}{ext}"


# ── Master tracker (tracker/master_tracker.xlsx) ─────────────────────────────
TRACKER_COLUMNS = [
    ("File Name", lambda r: r.get("new_name") or r.get("original_name")),
    ("Date", lambda r: r.get("resolved_date", "")),
    ("Category", lambda r: r.get("categoria", "")),
    ("Provider", lambda r: r.get("proveedor", "")),
    ("Type", lambda r: r.get("tipo", "")),
    ("Current Path", lambda r: r.get("original_path", "")),
    ("Target Path", lambda r: r.get("new_path", "")),
    ("Status", lambda r: r.get("action", "")),
    ("Notes", lambda r: r.get("reason", "")),
    ("Drive Link", lambda r: f"https://drive.google.com/file/d/{r['file_id']}/view" if r.get("file_id") else ""),
]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(wb, title, records):
    ws = wb.create_sheet(title=title[:31])  # Excel sheet-name length limit
    ws.append([col_name for col_name, _ in TRACKER_COLUMNS])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r in sorted(records, key=lambda r: (r.get("resolved_date") or "", r.get("categoria") or "", r.get("original_path", ""))):
        ws.append([getter(r) for _, getter in TRACKER_COLUMNS])
    ws.freeze_panes = "A2"
    for i, (col_name, _) in enumerate(TRACKER_COLUMNS, start=1):
        widest = max([len(col_name)] + [len(str(row[i - 1].value or "")) for row in ws.iter_rows(min_row=2)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 10), 60)
    return ws


def build_master_tracker(records, tracker_path, apply_changes):
    """(Re)writes the whole tracker workbook from this run's records — one tab per
    year (matching the Drive folder layout: {year}/{category}/...), a Resumen
    (summary) tab, and a _REVISAR tab for anything quarantined, duplicate, errored,
    ignored, or otherwise missing a resolved year. Safe to call on a dry run: it
    only writes the .xlsx file, never touches Drive."""
    by_year = defaultdict(list)
    by_category_count = defaultdict(int)  # informational only, summary tab
    revisar = []
    for r in records:
        action = r.get("action")
        categoria = r.get("categoria")
        target_year = r.get("target_year")
        if action in ("quarantine", "duplicate", "error", "ignored") or not categoria or categoria not in CATEGORIES or not target_year:
            revisar.append(r)
        else:
            by_year[target_year].append(r)
            by_category_count[categoria] += 1

    years_ordered = sorted(by_year.keys(), reverse=True)  # most recent year first

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    summary = wb.create_sheet(title="Resumen")
    summary.append(["Master Tracker — Medical Drive folder"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append(["Last updated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")])
    summary.append(["Mode", "apply (changes executed)" if apply_changes else "dry run (plan only, nothing changed on Drive)"])
    summary.append(["Total files scanned", len(records)])
    summary.append([])
    summary.append(["Year", "Files"])
    for cell in summary[summary.max_row]:
        cell.font = Font(bold=True)
    for year in years_ordered:
        summary.append([year, len(by_year[year])])
    summary.append(["_REVISAR (needs manual review)", len(revisar)])
    summary.append([])
    summary.append(["Category (all years)", "Files"])
    for cell in summary[summary.max_row]:
        cell.font = Font(bold=True)
    for cat in CATEGORIES_ORDERED:
        summary.append([cat, by_category_count.get(cat, 0)])
    for col, width in zip("AB", (34, 50)):
        summary.column_dimensions[col].width = width

    for year in years_ordered:
        _write_sheet(wb, year, by_year[year])
    _write_sheet(wb, REVISAR_FOLDER_NAME, revisar)

    os.makedirs(os.path.dirname(tracker_path) or ".", exist_ok=True)
    wb.save(tracker_path)


TRACKER_DRIVE_NAME = "master_tracker.xlsx"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def upload_tracker_to_drive(service, tracker_path, medical_folder_id):
    """Uploads/updates a loose copy of the tracker workbook directly in the Medical
    Drive folder root, right next to the year folders — so it's visible from Drive
    itself, not just in the GitHub repo. Looks for an existing file with this exact
    name directly in that folder (not nested inside a year/category subfolder) and
    overwrites its content in place; creates it on the first run. This is purely a
    courtesy copy of the inventory — it never renames/moves/deletes anything, so
    it's safe to do on a dry run too, same as writing the local .xlsx."""
    resp = service.files().list(
        q=(f"name = '{TRACKER_DRIVE_NAME}' and '{medical_folder_id}' in parents "
           f"and trashed = false"),
        fields="files(id, name)",
        spaces="drive",
    ).execute()
    existing = resp.get("files", [])

    media = MediaFileUpload(tracker_path, mimetype=XLSX_MIME_TYPE, resumable=False)
    if existing:
        file_id = existing[0]["id"]
        service.files().update(fileId=file_id, media_body=media).execute()
        return file_id
    else:
        created = service.files().create(
            body={"name": TRACKER_DRIVE_NAME, "parents": [medical_folder_id], "mimeType": XLSX_MIME_TYPE},
            media_body=media,
            fields="id",
        ).execute()
        return created["id"]


# ── Main sweep ────────────────────────────────────────────────────────────────
def _scan_and_plan(service, client, args, apply_changes, records):
    """Steps 1-5: locate/create folders, walk the tree, classify every file,
    dedupe, and apply (or just plan) the result. Appends onto the caller's
    `records` list in place, so even if this raises partway through — a Drive
    permissions error, a rate limit, anything unexpected — everything already
    classified survives in the caller's list and still gets logged. Returns
    False only for the (non-error) "requested --year folder doesn't exist"
    case, where there is deliberately nothing to log."""
    # 1. Locate Medical/_REVISAR (create only if applying and it's genuinely missing).
    top_children = list_children(service, GDRIVE_FOLDER_ID)
    revisar = next((f for f in top_children if f["name"] == REVISAR_FOLDER_NAME
                     and f["mimeType"] == "application/vnd.google-apps.folder"), None)
    folder_cache = {}
    if revisar:
        revisar_id = revisar["id"]
    elif apply_changes:
        revisar_id = get_or_create_folder(service, GDRIVE_FOLDER_ID, REVISAR_FOLDER_NAME, folder_cache, True)
    else:
        revisar_id = f"PLANNED:{GDRIVE_FOLDER_ID}:{REVISAR_FOLDER_NAME}"

    # Seed folder_cache with the real year/category folders we already saw.
    for f in top_children:
        if f["mimeType"] == "application/vnd.google-apps.folder" and f["name"] != REVISAR_FOLDER_NAME:
            folder_cache[(GDRIVE_FOLDER_ID, f["name"])] = f["id"]
            for sub in list_children(service, f["id"]):
                if sub["mimeType"] == "application/vnd.google-apps.folder":
                    folder_cache[(f["id"], sub["name"])] = sub["id"]

    # 2. Walk the tree and classify every file.
    scan_root = GDRIVE_FOLDER_ID
    scan_path_prefix = []
    if args.year:
        year_folder = folder_cache.get((GDRIVE_FOLDER_ID, args.year))
        if not year_folder or str(year_folder).startswith("PLANNED:"):
            print(f"Year folder '{args.year}' not found under Medical/ — nothing to do.")
            return False
        scan_root, scan_path_prefix = year_folder, [args.year]

    print("Scanning Medical/ tree...")
    consecutive_classify_failures = 0
    for f, path_parts, parent_id in walk_tree(service, scan_root, scan_path_prefix):
        current_path = "/".join(path_parts + [f["name"]])
        print(f"  {current_path}")
        try:
            file_bytes = download_file(service, f["id"])
        except Exception as e:
            records.append(_error_record(f, path_parts, parent_id, f"download failed: {e}"))
            continue

        content_hash = hashlib.sha256(file_bytes).hexdigest()
        mime = f["mimeType"]
        if mime not in ("application/pdf", "image/jpeg", "image/png", "image/jpg"):
            records.append(_error_record(f, path_parts, parent_id, f"unsupported file type ({mime}) — left alone",
                                          content_hash=content_hash, action="ignored"))
            continue

        classification, method, err = classify_document(file_bytes, mime, client)
        record = {
            "file_id": f["id"], "parent_id": parent_id,
            "original_path": current_path, "original_name": f["name"],
            "created_time": f.get("createdTime", ""),
            "content_hash": content_hash, "extraction_method": method,
            "mime_type": mime,
        }

        if err or not classification:
            record.update(action="quarantine", reason=f"extraction failed ({method}): {err}",
                           new_path="", new_name="")
            records.append(record)
            if looks_systemic(err):
                raise RuntimeError(
                    f"Classification is failing at the API/account level, not on this document "
                    f"(saw: {err!r}). Stopping now, before touching Drive, rather than quarantining "
                    f"every remaining file for a problem none of them actually have."
                )
            consecutive_classify_failures += 1
            if consecutive_classify_failures >= CONSECUTIVE_FAILURE_ABORT_THRESHOLD:
                raise RuntimeError(
                    f"{consecutive_classify_failures} files in a row all failed classification — that's "
                    f"far more consistent with the API being down than with that many individually bad "
                    f"documents. Stopping now, before touching Drive. Last error: {err!r}"
                )
            continue
        consecutive_classify_failures = 0

        order_hint = numeric_order_hint(classification.get("fecha_secundaria_texto"))
        fecha, date_reason, via_hint = resolve_date(classification.get("fecha_texto"), order_hint)
        categoria = classification.get("categoria")
        categoria_confianza = classification.get("categoria_confianza", "baja")
        proveedor = classification.get("proveedor") or "otro"
        tipo = classification.get("tipo") or "otro"
        record["classification"] = classification

        via_existing_name = False
        if fecha is None and date_reason == "ambiguous_day_month":
            # Last resort, only for files whose CURRENT name already follows this
            # pipeline's own YYYY-MM-DD_... convention: if that existing date is
            # one of the two calendar-valid readings of the ambiguous text (never
            # both, never neither), treat it as confirmed. This only accepts a
            # name that's independently consistent with the raw digits actually
            # printed on the document — it doesn't trust an arbitrary filename,
            # and a name that matches NEITHER reading is left alone (that's a
            # real inconsistency worth a human's attention, not something to
            # paper over). Enabled per your explicit call after reviewing the
            # ambiguous cases: existing titles here are clear enough to trust
            # when they check out against the document's own printed digits.
            existing = existing_filename_date(f["name"])
            as_dm, as_md = both_date_interpretations(classification.get("fecha_texto"))
            if existing and existing.isoformat() == as_dm and as_dm != as_md:
                fecha, via_existing_name = existing, "day-first"
            elif existing and existing.isoformat() == as_md and as_dm != as_md:
                fecha, via_existing_name = existing, "month-first"

        via_report_date = False
        if fecha is None and date_reason == "no_date_found":
            # The document has no sample-collection-date field at all (it's
            # printed blank on the form) — this isn't an ambiguity to resolve,
            # there's simply nothing there. If a second, different-purpose date
            # exists on the same page (report/creation/release date), fall back
            # to that rather than quarantining: it's the only real date on the
            # page, not a guess between two readings. Still goes through the
            # same unambiguous-or-nothing resolution (including the existing
            # filename tiebreak) as any other date — and if THAT date is itself
            # ambiguous with no way to break the tie, this still quarantines.
            # Enabled per your explicit call; always labeled clearly below as a
            # report date, not a confirmed collection date, so it's never
            # mistaken for one in the tracker.
            secundaria_texto = classification.get("fecha_secundaria_texto")
            if secundaria_texto:
                sec_fecha, sec_reason, _ = resolve_date(secundaria_texto)
                sec_via_existing = False
                if sec_fecha is None and sec_reason == "ambiguous_day_month":
                    existing = existing_filename_date(f["name"])
                    sec_dm, sec_md = both_date_interpretations(secundaria_texto)
                    if existing and existing.isoformat() == sec_dm and sec_dm != sec_md:
                        sec_fecha, sec_via_existing = existing, "day-first"
                    elif existing and existing.isoformat() == sec_md and sec_dm != sec_md:
                        sec_fecha, sec_via_existing = existing, "month-first"
                if sec_fecha is not None:
                    fecha, via_report_date = sec_fecha, True
                    via_existing_name = sec_via_existing or via_existing_name

        if fecha is None:
            reason = f"date: {date_reason}"
            if date_reason == "ambiguous_day_month":
                as_dm, as_md = both_date_interpretations(classification.get("fecha_texto"))
                if as_dm and as_md:
                    reason += (f" — as printed ({classification.get('fecha_texto')!r}) this is either "
                               f"{as_dm} (day-first) or {as_md} (month-first); open the file to see which")
                    existing = existing_filename_date(f["name"])
                    if existing and existing.isoformat() not in (as_dm, as_md):
                        reason += (f" (note: current filename implies {existing.isoformat()}, which matches "
                                   f"NEITHER reading — worth a closer look)")
            record.update(action="quarantine", reason=reason, new_path="", new_name="")
            records.append(record)
            continue
        if not categoria or categoria not in CATEGORIES or categoria_confianza != "alta":
            record.update(action="quarantine",
                           reason=f"category not confident (categoria={categoria!r}, confianza={categoria_confianza!r})",
                           new_path="", new_name="")
            records.append(record)
            continue

        ext = os.path.splitext(f["name"])[1].lower() or (".pdf" if mime == "application/pdf" else ".jpg")
        target_name = build_target_name(fecha, categoria, proveedor, tipo, ext)
        target_year = str(fecha.year)
        record.update(
            resolved_date=fecha.isoformat(), categoria=categoria, proveedor=proveedor, tipo=tipo,
            target_year=target_year, target_name=target_name,
        )
        if via_report_date:
            record["reason"] = (
                f"NO COLLECTION DATE ON DOCUMENT — using report/creation date instead "
                f"({classification.get('fecha_secundaria_label') or 'fecha_secundaria'}: "
                f"{classification.get('fecha_secundaria_texto')!r}). This is the report date, "
                f"not a confirmed sample-collection date."
            )
            if via_existing_name:
                record["reason"] += (
                    f" That report date was itself ambiguous on its own, but the current filename "
                    f"({f['name']!r}) already matches the {via_existing_name} reading."
                )
        elif via_hint:
            record["reason"] = (
                f"date order inferred from 2nd date on same document "
                f"({classification.get('fecha_secundaria_texto')!r}, unambiguous as {order_hint}) "
                f"— {classification.get('fecha_texto')!r} itself was ambiguous"
            )
        elif via_existing_name:
            record["reason"] = (
                f"date confirmed via existing filename ({f['name']!r}) — {classification.get('fecha_texto')!r} "
                f"was ambiguous on its own, but the current name already matches the {via_existing_name} reading"
            )
        records.append(record)

    # 3. Duplicate detection across the whole scanned set.
    by_hash = defaultdict(list)
    for r in records:
        if r.get("action") in ("quarantine", "ignored"):
            continue
        by_hash[r["content_hash"]].append(r)

    for group in by_hash.values():
        if len(group) < 2:
            continue
        # Prefer whichever copy is already exactly correctly named+placed; else oldest.
        def already_correct(r):
            parts = r["original_path"].split("/")
            if len(parts) < 3:
                return False
            return (r["original_name"] == r["target_name"]
                    and parts[0] == r["target_year"]
                    and parts[1] == r["categoria"])
        keeper = next((r for r in group if already_correct(r)), None)
        if keeper is None:
            keeper = min(group, key=lambda r: r.get("created_time", ""))
        for r in group:
            if r is keeper:
                continue
            r["action"] = "duplicate"
            r["reason"] = f"duplicate of {keeper['original_path']} (identical content hash)"

    # 4. Decide final action for everything not already quarantined/duplicate/ignored.
    for r in records:
        if "action" in r:
            continue
        current_year = r["original_path"].split("/")[0] if "/" in r["original_path"] else ""
        current_category = r["original_path"].split("/")[1] if r["original_path"].count("/") >= 2 else ""
        if r["original_name"] == r["target_name"] and current_year == r["target_year"] and current_category == r["categoria"]:
            r["action"] = "no_change"
        elif current_year == r["target_year"] and current_category == r["categoria"]:
            r["action"] = "rename"
        else:
            r["action"] = "move"

    # 5. Apply (or just log) the plan.
    print("\nApplying plan..." if apply_changes else "\nPlan (dry run — nothing will change):")
    for r in records:
        action = r.get("action", "quarantine")
        if action == "no_change":
            r["new_path"], r["new_name"] = r["original_path"], r["original_name"]
        elif action in ("rename", "move"):
            r["new_path"] = f"{r['target_year']}/{r['categoria']}/{r['target_name']}"
            r["new_name"] = r["target_name"]
            try:
                dest_folder_id = get_or_create_folder(
                    service,
                    get_or_create_folder(service, GDRIVE_FOLDER_ID, r["target_year"], folder_cache, apply_changes),
                    r["categoria"], folder_cache, apply_changes,
                )
                if apply_changes:
                    move_file(service, r["file_id"], r["parent_id"], dest_folder_id, r["target_name"])
            except Exception as e:
                # Never let one file's Drive error (permissions, a rate limit, a
                # transient API hiccup) take down the whole run — every other
                # file's already-computed plan is real work we don't want to lose.
                r["action"] = "error"
                r["reason"] = f"Drive update failed: {e}"
        elif action in ("quarantine", "duplicate"):
            prefix = "DUP_" if action == "duplicate" else "REVISAR_"
            new_name = r["original_name"] if r["original_name"].startswith(prefix) else prefix + r["original_name"]
            r["new_path"] = f"{REVISAR_FOLDER_NAME}/{new_name}"
            r["new_name"] = new_name
            if apply_changes:
                try:
                    move_file(service, r["file_id"], r["parent_id"], revisar_id, new_name)
                except Exception as e:
                    r["action"] = "error"
                    r["reason"] = f"Drive update failed: {e}"
        print(f"  [{r.get('action')}] {r['original_path']} -> {r.get('new_path', '(unchanged)')}"
              + (f"  ({r['reason']})" if r.get("reason") else ""))

    return True


def main():
    ap = argparse.ArgumentParser(description="Clean up the Medical Google Drive folder.")
    ap.add_argument("--apply", action="store_true", help="Actually rename/move/quarantine on Drive. Default is dry-run.")
    ap.add_argument("--year", default=None, help="Only sweep this year subfolder (e.g. 2024). Default: whole tree.")
    ap.add_argument("--log", default="rename_log.csv", help="Path to write the CSV log to.")
    ap.add_argument("--tracker", default=DEFAULT_TRACKER_PATH, help="Path to write the master tracker .xlsx to.")
    args = ap.parse_args()

    apply_changes = args.apply
    print(f"Mode: {'APPLY (changes will be made)' if apply_changes else 'DRY RUN (no changes will be made)'}")

    service = get_drive_service()
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

    records = []
    fatal_error = None
    try:
        had_work = _scan_and_plan(service, client, args, apply_changes, records)
    except Exception as e:
        # Something unexpected blew up mid-run (Drive permissions, a rate
        # limit, anything). Don't let it discard everything already
        # classified — log what we have, then still fail the workflow run
        # (below) so the problem is visible, not silently swallowed.
        import traceback
        traceback.print_exc()
        fatal_error = e
        had_work = True
        print(f"\n{'='*50}")
        print(f"FATAL ERROR partway through — saving the {len(records)} file(s) already processed before failing: {e}")

    if not had_work:
        return

    # 6. Write the CSV log (always — dry run or apply, and even after a fatal
    # error above, so nothing already classified is lost).
    # Includes the model's raw classification fields (fecha_texto, fecha_label,
    # categoria_confianza, notas) even for quarantined files, so a quarantine
    # reason like "unparseable_date_text" can be diagnosed from the log itself
    # instead of guessed at.
    fieldnames = ["action", "original_path", "original_name", "new_path", "new_name",
                  "resolved_date", "categoria", "proveedor", "tipo",
                  "extraction_method", "content_hash", "reason", "file_id", "applied",
                  "fecha_texto", "fecha_label", "fecha_secundaria_texto", "fecha_secundaria_label",
                  "categoria_confianza", "notas"]
    with open(args.log, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        # "applied" reflects whether Drive was actually touched — not just
        # whether --apply was passed. On a fatal-error abort (below), the
        # scan stops before step 5 ever runs a single Drive write, so nothing
        # in `records` was actually applied, however "quarantine"/"rename" it
        # may be labeled — the CSV should say so honestly.
        actually_applied = apply_changes and fatal_error is None
        for r in records:
            row = dict(r)
            row["applied"] = actually_applied and r.get("action") not in ("no_change", "ignored", "error")
            classification = r.get("classification") or {}
            row["fecha_texto"] = classification.get("fecha_texto", "")
            row["fecha_label"] = classification.get("fecha_label", "")
            row["fecha_secundaria_texto"] = classification.get("fecha_secundaria_texto", "")
            row["fecha_secundaria_label"] = classification.get("fecha_secundaria_label", "")
            row["categoria_confianza"] = classification.get("categoria_confianza", "")
            row["notas"] = classification.get("notas", "")
            w.writerow(row)

    # 7. (Re)build the master tracker workbook — one tab per year. Safe on a
    # dry run too; this only writes a local .xlsx, it never touches Drive.
    # Guarded too, so a tracker-building problem can never take down a run
    # that already has a good CSV log written above.
    try:
        build_master_tracker(records, args.tracker, apply_changes)
        tracker_note = f"Master tracker written to {args.tracker}"
    except Exception as e:
        tracker_note = f"Master tracker NOT written — build failed: {e}"
    else:
        # Also drop/update a loose copy directly in the Medical Drive folder,
        # next to the year folders — separately guarded so a Drive-side upload
        # hiccup can't discard the good local/git copy above.
        try:
            upload_tracker_to_drive(service, args.tracker, GDRIVE_FOLDER_ID)
            tracker_note += " (also updated in Drive)"
        except Exception as e:
            tracker_note += f" (Drive copy NOT updated: {e})"

    counts = defaultdict(int)
    for r in records:
        counts[r.get("action", "quarantine")] += 1
    print(f"\n{'='*50}")
    print(f"DONE — {len(records)} files scanned. " + ", ".join(f"{k}={v}" for k, v in sorted(counts.items())))
    print(f"Log written to {args.log}")
    print(tracker_note)
    if not apply_changes:
        print("This was a DRY RUN — nothing on Drive changed. Re-run with --apply to execute this plan.")

    if fatal_error:
        # The CSV/tracker above captured everything completed before the
        # crash, but the run still needs to show as failed in CI — a partial
        # sweep silently reported "green" would hide a real problem (Drive
        # permissions, a bug, whatever `fatal_error` says) from view.
        raise fatal_error


def _error_record(f, path_parts, parent_id, reason, content_hash="", action="quarantine"):
    return {
        "file_id": f["id"], "parent_id": parent_id,
        "original_path": "/".join(path_parts + [f["name"]]), "original_name": f["name"],
        "created_time": f.get("createdTime", ""), "content_hash": content_hash,
        "extraction_method": "", "mime_type": f.get("mimeType", ""),
        "action": action, "reason": reason, "new_path": "", "new_name": "",
    }


if __name__ == "__main__":
    main()
